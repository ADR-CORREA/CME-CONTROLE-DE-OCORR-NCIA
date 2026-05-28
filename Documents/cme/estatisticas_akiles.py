#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
estatisticas_akiles.py
======================
Gerador de relatório estatístico do sistema Akiles — Setor de Monitoramento CME/IAPEN-AP.

Uso:
    python estatisticas_akiles.py <arquivo_akiles.xls> [saida.xlsx]

Dependências:
    pip install pandas openpyxl lxml

Gera um arquivo Excel com 12 abas padronizadas, gráficos e formatação ABNT/institucional.
"""

import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import os
import re
from datetime import datetime

try:
    import pandas as pd
except ImportError:
    sys.exit("❌  pandas não encontrado. Execute: pip install pandas openpyxl lxml")

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.chart.label import DataLabel
    from openpyxl.drawing.fill import PatternFillProperties
    from openpyxl.chart.data_source import NumDataSource, NumRef
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.drawing.spreadsheet_drawing import SpreadsheetDrawing
except ImportError:
    sys.exit("❌  openpyxl não encontrado. Execute: pip install openpyxl")

# ── Paleta de cores ──────────────────────────────────────────────────────────
AZUL_HEADER    = "1F4E78"   # cabeçalho principal
AZUL_SUBHEADER = "2E75B6"   # subcabeçalho
AZUL_CLARO     = "DAEEF3"   # fundo zebrastripe par
BRANCO         = "FFFFFF"
CINZA_BORDA    = "B8CCE4"

# Série: 6 colunas (Por Tipo, Sexo, Município, Faixa Etária)
COR_MULHER_DPP      = "ED7D31"   # Mulher Vítima DPP
COR_MULHER_PRIS     = "F4B084"   # Mulher Sistema Prisional
COR_HOMEM_DPP       = "A6A6A6"   # Homem Vítima DPP
COR_HOMEM_PRIS      = "595959"   # Homem Sistema Prisional
COR_TOTAL           = "1F4E78"   # Total Geral

# Header 4 colunas (Por Grupo, Vara, Operador, Periculosidade)
COR_HDR_DPP         = "F8CBAD"   # cabeçalho Vítima DPP
COR_HDR_PRIS        = "C9C9C9"   # cabeçalho Sistema Prisional

# ── Estilos base ─────────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=11, color="000000", name="Arial"):
    return Font(name=name, bold=bold, size=size, color=color)

def _border_thin():
    s = Side(style="thin", color=CINZA_BORDA)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ── Helpers de escrita ────────────────────────────────────────────────────────
def _write(ws, row, col, value, bold=False, size=11, fill_hex=None,
           color="000000", h_align="center", wrap=True, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = _font(bold=bold, size=size, color=color)
    cell.alignment = _align(h=h_align, wrap=wrap)
    cell.border    = _border_thin()
    if fill_hex:
        cell.fill = _fill(fill_hex)
    if number_format:
        cell.number_format = number_format
    return cell

def _header_row(ws, row, cols_labels, fills, bold=True, size=11, font_color="FFFFFF"):
    """Escreve uma linha de cabeçalho com lista de (label, fill_hex)."""
    for col_idx, (label, fill_hex) in enumerate(zip(cols_labels, fills), start=1):
        _write(ws, row, col_idx, label, bold=bold, size=size,
               fill_hex=fill_hex, color=font_color)

def _set_col_widths(ws, widths):
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

def _freeze(ws, cell="A2"):
    ws.freeze_panes = cell

# ── Leitura do arquivo ────────────────────────────────────────────────────────
def ler_arquivo(path):
    """Lê o arquivo Akiles (HTML disfarçado de .xls ou .xlsx real) e retorna DataFrame."""
    ext = os.path.splitext(path)[1].lower()
    df = None
    # 1ª tentativa: pd.read_html (formato HTML disfarçado de .xls)
    try:
        tabelas = pd.read_html(path, encoding="utf-8", header=0)
        if not tabelas:
            tabelas = pd.read_html(path, encoding="latin-1", header=0)
        df = max(tabelas, key=lambda t: len(t))
        print(f"   ✔ Lido via HTML ({len(df)} linhas, {len(df.columns)} colunas)")
        return df
    except Exception:
        pass

    # 2ª tentativa: openpyxl / xlrd
    try:
        df = pd.read_excel(path, engine="openpyxl" if ext == ".xlsx" else None)
        print(f"   ✔ Lido via Excel ({len(df)} linhas, {len(df.columns)} colunas)")
        return df
    except Exception:
        pass

    sys.exit(f"❌  Não foi possível ler o arquivo: {path}\n"
             "    Verifique se está no formato correto (exportação Akiles).")

# ── Normalização ──────────────────────────────────────────────────────────────
def normalizar(df):
    """Limpa e padroniza colunas essenciais."""
    # Padronizar nomes de colunas (strip, title-case)
    df.columns = [str(c).strip() for c in df.columns]

    # Mapa de colunas esperadas → possíveis nomes no Akiles
    mapa = {
        "Nome":               ["Nome", "Nome Completo", "Monitorado"],
        "Sexo":               ["Sexo", "Genero", "Gênero"],
        "Tipo":               ["Tipo", "Tipo Monitorado", "Categoria Monitorado"],
        "Municipio":          ["Município", "Municipio", "Cidade"],
        "Vara":               ["Vara", "Vara/Juízo", "Vara / Juízo", "Juízo", "Juizo"],
        "Grupo":              ["Grupo", "Grupo/Regime", "Regime", "Grupo / Regime"],
        "FaixaEtaria":        ["Faixa Etária", "Faixa Etaria", "Faixa", "Idade"],
        "Operador":           ["Operador", "Fiscal", "Técnico", "Técnico Responsável"],
        "Periculosidade":     ["Periculosidade", "Risco", "Nível de Risco", "Classificação"],
        "ConfigDispositivo":  ["Config Dispositivo", "Configuração Dispositivo",
                               "Config. Dispositivo", "Status Dispositivo",
                               "Dispositivo", "Config"],
        "MedidaProtetiva":    ["Medida Protetiva", "Medida", "MP", "Tipo Medida"],
        "StatusDispositivo":  ["Status Dispositivo", "Status", "Situação Dispositivo"],
    }

    # Renomear colunas encontradas
    rename = {}
    for chave, candidatos in mapa.items():
        for c in candidatos:
            if c in df.columns and chave not in rename.values():
                rename[c] = chave
                break

    df = df.rename(columns=rename)

    # Garantir colunas obrigatórias
    for col in ["Nome", "Sexo", "Tipo", "ConfigDispositivo"]:
        if col not in df.columns:
            df[col] = ""

    # Normalizar strings
    for col in df.select_dtypes(include="object").columns:
        df[col] = df[col].fillna("").astype(str).str.strip()

    # Categoria: Vítima (DPP) vs Sistema Prisional
    df["Categoria"] = df["Tipo"].apply(
        lambda t: "Vítima (DPP)" if re.search(r"v[ií]tima", t, re.I) else "Sistema Prisional"
    )

    # Sexo normalizado
    df["SexoNorm"] = df["Sexo"].apply(
        lambda s: "Feminino" if re.search(r"f(em|\.)?", s, re.I) else "Masculino"
    )

    return df

def filtrar_ativos(df):
    """Retorna apenas monitorados ativos/pré-ativos."""
    if "ConfigDispositivo" not in df.columns:
        return df
    ativos = ["ativo", "pre-ativo", "pré-ativo", "pre ativo"]
    mask = df["ConfigDispositivo"].str.lower().str.strip().isin(ativos)
    return df[mask].copy()

# ── Pivô auxiliar ─────────────────────────────────────────────────────────────
def _pivot_6col(df_ativos, col_agrupamento):
    """
    Cria tabela com 6 colunas:
    col_agrupamento | Mulher DPP | Mulher Pris | Homem DPP | Homem Pris | Total
    Separado por Categoria (DPP / Prisional) com subtotal.
    """
    if col_agrupamento not in df_ativos.columns:
        return pd.DataFrame()

    registros = []
    for cat in ["Vítima (DPP)", "Sistema Prisional"]:
        sub = df_ativos[df_ativos["Categoria"] == cat]
        grupos = sub[col_agrupamento].unique()
        for g in sorted(grupos):
            g_df = sub[sub[col_agrupamento] == g]
            m_dpp  = len(g_df[(g_df["SexoNorm"] == "Feminino")]) if cat == "Vítima (DPP)" else 0
            h_dpp  = len(g_df[(g_df["SexoNorm"] == "Masculino")]) if cat == "Vítima (DPP)" else 0
            m_pris = len(g_df[(g_df["SexoNorm"] == "Feminino")]) if cat == "Sistema Prisional" else 0
            h_pris = len(g_df[(g_df["SexoNorm"] == "Masculino")]) if cat == "Sistema Prisional" else 0
            total  = len(g_df)
            registros.append({
                "Categoria":      cat,
                col_agrupamento:  g,
                "Mulher DPP":     m_dpp,
                "Mulher Pris":    m_pris,
                "Homem DPP":      h_dpp,
                "Homem Pris":     h_pris,
                "Total Geral":    total,
            })
        # Subtotal por categoria
        sub_total = len(sub)
        registros.append({
            "Categoria":      f"SUBTOTAL {cat.upper()}",
            col_agrupamento:  "",
            "Mulher DPP":     len(sub[sub["SexoNorm"] == "Feminino"]) if cat == "Vítima (DPP)" else 0,
            "Mulher Pris":    len(sub[sub["SexoNorm"] == "Feminino"]) if cat == "Sistema Prisional" else 0,
            "Homem DPP":      len(sub[sub["SexoNorm"] == "Masculino"]) if cat == "Vítima (DPP)" else 0,
            "Homem Pris":     len(sub[sub["SexoNorm"] == "Masculino"]) if cat == "Sistema Prisional" else 0,
            "Total Geral":    sub_total,
        })

    # Linha TOTAL GERAL
    registros.append({
        "Categoria":     "TOTAL GERAL",
        col_agrupamento: "",
        "Mulher DPP":    len(df_ativos[(df_ativos["Categoria"] == "Vítima (DPP)") & (df_ativos["SexoNorm"] == "Feminino")]),
        "Mulher Pris":   len(df_ativos[(df_ativos["Categoria"] == "Sistema Prisional") & (df_ativos["SexoNorm"] == "Feminino")]),
        "Homem DPP":     len(df_ativos[(df_ativos["Categoria"] == "Vítima (DPP)") & (df_ativos["SexoNorm"] == "Masculino")]),
        "Homem Pris":    len(df_ativos[(df_ativos["Categoria"] == "Sistema Prisional") & (df_ativos["SexoNorm"] == "Masculino")]),
        "Total Geral":   len(df_ativos),
    })

    return pd.DataFrame(registros)

def _pivot_4col(df_ativos, col_agrupamento):
    """
    Cria tabela com 4 colunas:
    col_agrupamento | Vítima DPP | Sistema Prisional | Total Geral
    """
    if col_agrupamento not in df_ativos.columns:
        return pd.DataFrame()

    grupos = sorted(df_ativos[col_agrupamento].unique())
    registros = []
    for g in grupos:
        g_df = df_ativos[df_ativos[col_agrupamento] == g]
        dpp  = len(g_df[g_df["Categoria"] == "Vítima (DPP)"])
        pris = len(g_df[g_df["Categoria"] == "Sistema Prisional"])
        registros.append({
            col_agrupamento:    g,
            "Vítima (DPP)":     dpp,
            "Sistema Prisional": pris,
            "Total Geral":      len(g_df),
        })

    # Total
    registros.append({
        col_agrupamento:    "TOTAL GERAL",
        "Vítima (DPP)":     len(df_ativos[df_ativos["Categoria"] == "Vítima (DPP)"]),
        "Sistema Prisional": len(df_ativos[df_ativos["Categoria"] == "Sistema Prisional"]),
        "Total Geral":      len(df_ativos),
    })

    return pd.DataFrame(registros)

# ── Título de aba ─────────────────────────────────────────────────────────────
def _titulo_aba(ws, titulo, subtitulo, data_ref):
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = titulo
    c.font      = _font(bold=True, size=14, color="FFFFFF")
    c.fill      = _fill(AZUL_HEADER)
    c.alignment = _align(h="center")

    ws.merge_cells("A2:H2")
    c2 = ws["A2"]
    c2.value     = subtitulo
    c2.font      = _font(bold=False, size=10, color="FFFFFF")
    c2.fill      = _fill(AZUL_SUBHEADER)
    c2.alignment = _align(h="center")

    ws.merge_cells("A3:H3")
    c3 = ws["A3"]
    c3.value     = f"Gerado em {data_ref}  |  Fonte: Sistema Akiles — CME/IAPEN-AP"
    c3.font      = _font(bold=False, size=9, color="595959")
    c3.alignment = _align(h="right")
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 14

# ── Escrita de tabela 6-col ───────────────────────────────────────────────────
COLS_6 = ["Categoria", "Valor",
          "Mulher\nVítima DPP", "Mulher\nSist. Prisional",
          "Homem\nVítima DPP", "Homem\nSist. Prisional",
          "Total\nGeral"]
FILLS_6_HDR = [AZUL_HEADER, AZUL_HEADER,
               COR_MULHER_DPP, COR_MULHER_PRIS,
               COR_HOMEM_DPP, COR_HOMEM_PRIS,
               AZUL_HEADER]

def _escrever_tabela_6col(ws, df_pivot, col_nome, row_start=5):
    """Escreve tabela 6 colunas a partir de row_start."""
    # Cabeçalho
    labels = ["Categoria", col_nome,
              "Mulher\nVítima DPP", "Mulher\nSist. Prisional",
              "Homem\nVítima DPP", "Homem\nSist. Prisional",
              "Total\nGeral"]
    fills_fnt = [("FFFFFF", AZUL_HEADER)] * 2 + \
                [("FFFFFF", COR_MULHER_DPP), ("000000", COR_MULHER_PRIS),
                 ("FFFFFF", COR_HOMEM_DPP), ("FFFFFF", COR_HOMEM_PRIS),
                 ("FFFFFF", AZUL_HEADER)]

    for ci, (lbl, (fcor, bcor)) in enumerate(zip(labels, fills_fnt), start=1):
        _write(ws, row_start, ci, lbl, bold=True, size=10,
               fill_hex=bcor, color=fcor)
    ws.row_dimensions[row_start].height = 28

    col_keys = ["Categoria", col_nome,
                "Mulher DPP", "Mulher Pris",
                "Homem DPP", "Homem Pris",
                "Total Geral"]

    for ri, (_, row) in enumerate(df_pivot.iterrows(), start=row_start + 1):
        eh_subtotal = str(row.get("Categoria", "")).startswith("SUBTOTAL")
        eh_total    = str(row.get("Categoria", "")) == "TOTAL GERAL"
        fill_bg     = AZUL_CLARO if ri % 2 == 0 else BRANCO
        if eh_subtotal:
            fill_bg = "FFF2CC"   # amarelo claro
        if eh_total:
            fill_bg = "D6E4F0"   # azul bem claro

        for ci, key in enumerate(col_keys, start=1):
            val = row.get(key, "")
            bold = eh_subtotal or eh_total
            _write(ws, ri, ci, val,
                   bold=bold, size=10, fill_hex=fill_bg,
                   h_align="center" if ci > 2 else "left")
        ws.row_dimensions[ri].height = 16

    return row_start + 1 + len(df_pivot)

def _escrever_tabela_4col(ws, df_pivot, col_nome, row_start=5):
    """Escreve tabela 4 colunas (Grupo, Vara, Operador, Periculosidade)."""
    labels = [col_nome, "Vítima (DPP)", "Sistema Prisional", "Total Geral"]
    fills_hdr = [AZUL_HEADER, COR_HDR_DPP, COR_HDR_PRIS, AZUL_HEADER]
    font_col  = ["FFFFFF", "000000", "000000", "FFFFFF"]

    for ci, (lbl, bcor, fcor) in enumerate(zip(labels, fills_hdr, font_col), start=1):
        _write(ws, row_start, ci, lbl, bold=True, size=10,
               fill_hex=bcor, color=fcor)
    ws.row_dimensions[row_start].height = 20

    col_keys = [col_nome, "Vítima (DPP)", "Sistema Prisional", "Total Geral"]

    for ri, (_, row) in enumerate(df_pivot.iterrows(), start=row_start + 1):
        eh_total = str(row.get(col_nome, "")) == "TOTAL GERAL"
        fill_bg  = AZUL_CLARO if ri % 2 == 0 else BRANCO
        if eh_total:
            fill_bg = "D6E4F0"

        for ci, key in enumerate(col_keys, start=1):
            val = row.get(key, "")
            _write(ws, ri, ci, val,
                   bold=eh_total, size=10, fill_hex=fill_bg,
                   h_align="center" if ci > 1 else "left")
        ws.row_dimensions[ri].height = 15

    return row_start + 1 + len(df_pivot)

# ── Gráfico de barras empilhadas ──────────────────────────────────────────────
def _add_bar_chart(ws, title, data_ref_start, data_ref_end, cat_ref_start,
                   cat_ref_end, anchor="I5", grouped=True):
    """Adiciona BarChart simples ao worksheet."""
    chart = BarChart()
    chart.type    = "col"
    chart.grouping = "clustered" if not grouped else "stacked"
    chart.title   = title
    chart.y_axis.title = "Quantidade"
    chart.x_axis.title = ""
    chart.style  = 10
    chart.width  = 18
    chart.height = 12

    ws.add_chart(chart, anchor)
    return chart

# ── ABA 1: Resumo ─────────────────────────────────────────────────────────────
def aba_resumo(wb, df, df_ativos, data_ref):
    ws = wb.create_sheet("Resumo")
    _titulo_aba(ws, "RESUMO EXECUTIVO — MONITORAMENTO ELETRÔNICO",
                "Visão consolidada de todos os indicadores do período", data_ref)
    _freeze(ws, "A5")

    total_geral    = len(df)
    total_ativos   = len(df_ativos)
    total_inativos = total_geral - total_ativos
    dpp   = len(df_ativos[df_ativos["Categoria"] == "Vítima (DPP)"])
    pris  = len(df_ativos[df_ativos["Categoria"] == "Sistema Prisional"])
    fem   = len(df_ativos[df_ativos["SexoNorm"] == "Feminino"])
    masc  = len(df_ativos[df_ativos["SexoNorm"] == "Masculino"])

    row = 5
    _write(ws, row, 1, "INDICADOR",      bold=True, fill_hex=AZUL_HEADER, color="FFFFFF", size=11)
    _write(ws, row, 2, "VALOR",          bold=True, fill_hex=AZUL_HEADER, color="FFFFFF", size=11)
    _write(ws, row, 3, "OBS",            bold=True, fill_hex=AZUL_HEADER, color="FFFFFF", size=11)

    indicadores = [
        ("Total de registros (base completa)",   total_geral,  "Todos os status"),
        ("Total Ativos / Pré-ativos",            total_ativos, "Base para análise"),
        ("Inativos / Outros",                    total_inativos, "Excluídos da análise"),
        ("",                                     "",           ""),
        ("Vítima (DPP) — ativos",                dpp,          f"{dpp/total_ativos*100:.1f}% do total ativo" if total_ativos else ""),
        ("Sistema Prisional — ativos",           pris,         f"{pris/total_ativos*100:.1f}% do total ativo" if total_ativos else ""),
        ("",                                     "",           ""),
        ("Sexo Feminino — ativos",               fem,          f"{fem/total_ativos*100:.1f}% do total ativo" if total_ativos else ""),
        ("Sexo Masculino — ativos",              masc,         f"{masc/total_ativos*100:.1f}% do total ativo" if total_ativos else ""),
    ]

    for i, (ind, val, obs) in enumerate(indicadores, start=row + 1):
        fill = AZUL_CLARO if i % 2 == 0 else BRANCO
        if not ind:
            fill = BRANCO
        _write(ws, i, 1, ind,  fill_hex=fill, h_align="left", size=10)
        _write(ws, i, 2, val,  fill_hex=fill, size=10)
        _write(ws, i, 3, obs,  fill_hex=fill, h_align="left", size=10)
        ws.row_dimensions[i].height = 16

    _set_col_widths(ws, [35, 12, 40])

    # Mini gráfico pizza DPP vs Prisional
    if total_ativos:
        row_chart = row + len(indicadores) + 3
        ws.cell(row=row_chart,   column=1, value="Categoria")
        ws.cell(row=row_chart+1, column=1, value="Vítima (DPP)")
        ws.cell(row=row_chart+2, column=1, value="Sistema Prisional")
        ws.cell(row=row_chart,   column=2, value="Qtd")
        ws.cell(row=row_chart+1, column=2, value=dpp)
        ws.cell(row=row_chart+2, column=2, value=pris)

        pie = PieChart()
        pie.title  = "Distribuição por Categoria"
        pie.style  = 10
        pie.width  = 14
        pie.height = 10
        data = Reference(ws, min_col=2, min_row=row_chart+1, max_row=row_chart+2)
        cats = Reference(ws, min_col=1, min_row=row_chart+1, max_row=row_chart+2)
        pie.add_data(data)
        pie.set_categories(cats)
        ws.add_chart(pie, "E5")

# ── ABA 2: Panorama Geral ─────────────────────────────────────────────────────
def aba_panorama(wb, df, df_ativos, data_ref):
    ws = wb.create_sheet("Panorama Geral")
    _titulo_aba(ws, "PANORAMA GERAL DO MONITORAMENTO",
                "Contagens globais por categoria, sexo e situação do dispositivo", data_ref)

    # Tabela 1: por Categoria
    row = 5
    _write(ws, row, 1, "Por Categoria (ativos)", bold=True, fill_hex=AZUL_SUBHEADER,
           color="FFFFFF", size=10)
    ws.merge_cells(f"A{row}:D{row}")
    row += 1
    for cat in ["Vítima (DPP)", "Sistema Prisional"]:
        n = len(df_ativos[df_ativos["Categoria"] == cat])
        fill = AZUL_CLARO if row % 2 == 0 else BRANCO
        _write(ws, row, 1, cat, fill_hex=fill, h_align="left", size=10)
        _write(ws, row, 2, n,   fill_hex=fill, size=10)
        pct = f"{n/len(df_ativos)*100:.1f}%" if len(df_ativos) else "0%"
        _write(ws, row, 3, pct, fill_hex=fill, size=10)
        row += 1
    fill = "D6E4F0"
    _write(ws, row, 1, "TOTAL GERAL", bold=True, fill_hex=fill, h_align="left", size=10)
    _write(ws, row, 2, len(df_ativos), bold=True, fill_hex=fill, size=10)
    _write(ws, row, 3, "100%", bold=True, fill_hex=fill, size=10)
    row += 2

    # Tabela 2: por Sexo
    _write(ws, row, 1, "Por Sexo (ativos)", bold=True, fill_hex=AZUL_SUBHEADER,
           color="FFFFFF", size=10)
    ws.merge_cells(f"A{row}:D{row}")
    row += 1
    for sexo in ["Feminino", "Masculino"]:
        n = len(df_ativos[df_ativos["SexoNorm"] == sexo])
        fill = AZUL_CLARO if row % 2 == 0 else BRANCO
        _write(ws, row, 1, sexo, fill_hex=fill, h_align="left", size=10)
        _write(ws, row, 2, n,    fill_hex=fill, size=10)
        pct = f"{n/len(df_ativos)*100:.1f}%" if len(df_ativos) else "0%"
        _write(ws, row, 3, pct,  fill_hex=fill, size=10)
        row += 1
    fill = "D6E4F0"
    _write(ws, row, 1, "TOTAL GERAL", bold=True, fill_hex=fill, h_align="left", size=10)
    _write(ws, row, 2, len(df_ativos), bold=True, fill_hex=fill, size=10)
    _write(ws, row, 3, "100%", bold=True, fill_hex=fill, size=10)
    row += 2

    # Tabela 3: por Config Dispositivo (todos os status da base completa)
    if "ConfigDispositivo" in df.columns:
        _write(ws, row, 1, "Por Config. Dispositivo (base completa)", bold=True,
               fill_hex=AZUL_SUBHEADER, color="FFFFFF", size=10)
        ws.merge_cells(f"A{row}:D{row}")
        row += 1
        counts = df["ConfigDispositivo"].value_counts()
        for status, n in counts.items():
            fill = AZUL_CLARO if row % 2 == 0 else BRANCO
            _write(ws, row, 1, status, fill_hex=fill, h_align="left", size=10)
            _write(ws, row, 2, n,      fill_hex=fill, size=10)
            pct = f"{n/len(df)*100:.1f}%" if len(df) else "0%"
            _write(ws, row, 3, pct,    fill_hex=fill, size=10)
            row += 1
        fill = "D6E4F0"
        _write(ws, row, 1, "TOTAL GERAL", bold=True, fill_hex=fill, h_align="left", size=10)
        _write(ws, row, 2, len(df), bold=True, fill_hex=fill, size=10)
        _write(ws, row, 3, "100%", bold=True, fill_hex=fill, size=10)

    _set_col_widths(ws, [35, 12, 12])
    _freeze(ws, "A5")

# ── ABA 3: Medidas Protetivas ─────────────────────────────────────────────────
def aba_medidas(wb, df_ativos, data_ref):
    ws = wb.create_sheet("Medidas Protetivas")
    _titulo_aba(ws, "MEDIDAS PROTETIVAS — MONITORADOS ATIVOS",
                "Distribuição por tipo de medida protetiva aplicada", data_ref)
    _freeze(ws, "A5")

    if "MedidaProtetiva" not in df_ativos.columns:
        ws["A5"] = "⚠ Coluna 'Medida Protetiva' não encontrada no arquivo."
        _set_col_widths(ws, [50])
        return

    df_pivot = _pivot_4col(df_ativos, "MedidaProtetiva")
    _escrever_tabela_4col(ws, df_pivot, "MedidaProtetiva", row_start=5)
    _set_col_widths(ws, [42, 18, 18, 14])

# ── ABA 4: Por Tipo ───────────────────────────────────────────────────────────
def aba_por_tipo(wb, df_ativos, data_ref):
    ws = wb.create_sheet("Por Tipo")
    _titulo_aba(ws, "MONITORADOS ATIVOS — POR TIPO",
                "Distribuição por tipo de monitoramento, categoria e sexo", data_ref)
    _freeze(ws, "A5")

    if "Tipo" not in df_ativos.columns:
        ws["A5"] = "⚠ Coluna 'Tipo' não encontrada."
        return

    df_pivot = _pivot_6col(df_ativos, "Tipo")
    last_row = _escrever_tabela_6col(ws, df_pivot, "Tipo", row_start=5)
    _set_col_widths(ws, [28, 30, 16, 18, 14, 16, 14])

# ── ABA 5: Por Sexo ───────────────────────────────────────────────────────────
def aba_por_sexo(wb, df_ativos, data_ref):
    ws = wb.create_sheet("Por Sexo")
    _titulo_aba(ws, "MONITORADOS ATIVOS — POR SEXO",
                "Distribuição por sexo e categoria de monitoramento", data_ref)
    _freeze(ws, "A5")

    df_pivot = _pivot_6col(df_ativos, "SexoNorm")
    _escrever_tabela_6col(ws, df_pivot, "Sexo", row_start=5)
    _set_col_widths(ws, [28, 14, 16, 18, 14, 16, 14])

# ── ABA 6: Por Município ──────────────────────────────────────────────────────
def aba_por_municipio(wb, df_ativos, data_ref):
    ws = wb.create_sheet("Por Municipio")
    _titulo_aba(ws, "MONITORADOS ATIVOS — POR MUNICÍPIO",
                "Distribuição geográfica por município de residência/vinculação", data_ref)
    _freeze(ws, "A5")

    if "Municipio" not in df_ativos.columns:
        ws["A5"] = "⚠ Coluna 'Município' não encontrada."
        return

    df_pivot = _pivot_6col(df_ativos, "Municipio")
    _escrever_tabela_6col(ws, df_pivot, "Município", row_start=5)
    _set_col_widths(ws, [28, 30, 16, 18, 14, 16, 14])

# ── ABA 7: Faixa Etária ───────────────────────────────────────────────────────
def aba_faixa_etaria(wb, df_ativos, data_ref):
    ws = wb.create_sheet("Faixa Etaria")
    _titulo_aba(ws, "MONITORADOS ATIVOS — POR FAIXA ETÁRIA",
                "Distribuição por faixa etária, categoria e sexo", data_ref)
    _freeze(ws, "A5")

    if "FaixaEtaria" not in df_ativos.columns:
        ws["A5"] = "⚠ Coluna 'Faixa Etária' não encontrada."
        return

    df_pivot = _pivot_6col(df_ativos, "FaixaEtaria")
    _escrever_tabela_6col(ws, df_pivot, "Faixa Etária", row_start=5)
    _set_col_widths(ws, [28, 22, 16, 18, 14, 16, 14])

# ── ABA 8: Por Grupo ──────────────────────────────────────────────────────────
def aba_por_grupo(wb, df_ativos, data_ref):
    ws = wb.create_sheet("Por Grupo")
    _titulo_aba(ws, "MONITORADOS ATIVOS — POR GRUPO/REGIME",
                "Distribuição por grupo de pena / regime prisional e categoria", data_ref)
    _freeze(ws, "A5")

    if "Grupo" not in df_ativos.columns:
        ws["A5"] = "⚠ Coluna 'Grupo/Regime' não encontrada."
        return

    df_pivot = _pivot_4col(df_ativos, "Grupo")
    _escrever_tabela_4col(ws, df_pivot, "Grupo", row_start=5)
    _set_col_widths(ws, [38, 18, 20, 14])

    # Gráfico de barras
    n_rows = len(df_pivot)
    data_ref_chart = Reference(ws, min_col=2, max_col=4, min_row=5, max_row=5 + n_rows)
    cats_ref = Reference(ws, min_col=1, min_row=6, max_row=5 + n_rows)
    chart = BarChart()
    chart.type     = "col"
    chart.grouping = "clustered"
    chart.title    = "Monitorados por Grupo/Regime"
    chart.style    = 10
    chart.width    = 20
    chart.height   = 12
    chart.add_data(data_ref_chart, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, "F5")

# ── ABA 9: Por Vara/Juízo ─────────────────────────────────────────────────────
def aba_por_vara(wb, df_ativos, data_ref):
    ws = wb.create_sheet("Por Vara-Juizo")
    _titulo_aba(ws, "MONITORADOS ATIVOS — POR VARA / JUÍZO",
                "Distribuição por vara ou juízo responsável pelo monitoramento", data_ref)
    _freeze(ws, "A5")

    if "Vara" not in df_ativos.columns:
        ws["A5"] = "⚠ Coluna 'Vara/Juízo' não encontrada."
        return

    df_pivot = _pivot_4col(df_ativos, "Vara")
    _escrever_tabela_4col(ws, df_pivot, "Vara", row_start=5)
    _set_col_widths(ws, [50, 18, 20, 14])

# ── ABA 10: Por Operador ──────────────────────────────────────────────────────
def aba_por_operador(wb, df_ativos, data_ref):
    ws = wb.create_sheet("Por Operador")
    _titulo_aba(ws, "MONITORADOS ATIVOS — POR OPERADOR / FISCAL",
                "Distribuição por técnico/fiscal responsável pelo caso", data_ref)
    _freeze(ws, "A5")

    if "Operador" not in df_ativos.columns:
        ws["A5"] = "⚠ Coluna 'Operador' não encontrada."
        return

    df_pivot = _pivot_4col(df_ativos, "Operador")
    _escrever_tabela_4col(ws, df_pivot, "Operador", row_start=5)
    _set_col_widths(ws, [40, 18, 20, 14])

# ── ABA 11: Por Periculosidade ────────────────────────────────────────────────
def aba_por_periculosidade(wb, df_ativos, data_ref):
    ws = wb.create_sheet("Por Periculosidade")
    _titulo_aba(ws, "MONITORADOS ATIVOS — POR PERICULOSIDADE",
                "Distribuição por nível de risco / periculosidade", data_ref)
    _freeze(ws, "A5")

    if "Periculosidade" not in df_ativos.columns:
        ws["A5"] = "⚠ Coluna 'Periculosidade' não encontrada."
        return

    df_pivot = _pivot_4col(df_ativos, "Periculosidade")
    _escrever_tabela_4col(ws, df_pivot, "Periculosidade", row_start=5)
    _set_col_widths(ws, [32, 18, 20, 14])

    # Gráfico pizza
    n_rows = len(df_pivot) - 1  # excluir TOTAL GERAL
    if n_rows > 0:
        data_ref_chart = Reference(ws, min_col=4, min_row=5, max_row=5 + n_rows)
        cats_ref = Reference(ws, min_col=1, min_row=6, max_row=5 + n_rows)
        pie = PieChart()
        pie.title  = "Distribuição por Periculosidade"
        pie.style  = 10
        pie.width  = 16
        pie.height = 10
        pie.add_data(data_ref_chart, titles_from_data=True)
        pie.set_categories(cats_ref)
        ws.add_chart(pie, "F5")

# ── ABA 12: Status Dispositivos ───────────────────────────────────────────────
def aba_status_dispositivos(wb, df, data_ref):
    ws = wb.create_sheet("Status Dispositivos")
    _titulo_aba(ws, "STATUS DOS DISPOSITIVOS — BASE COMPLETA",
                "Situação dos dispositivos eletrônicos de todos os monitorados (base completa)", data_ref)
    _freeze(ws, "A5")

    if "ConfigDispositivo" not in df.columns:
        ws["A5"] = "⚠ Coluna 'Config Dispositivo' não encontrada."
        _set_col_widths(ws, [50])
        return

    row = 5
    # Cabeçalho
    hdrs = ["Config. Dispositivo", "Vítima (DPP)", "Sistema Prisional", "Total Geral", "%"]
    fills_hdr = [AZUL_HEADER, COR_HDR_DPP, COR_HDR_PRIS, AZUL_HEADER, AZUL_HEADER]
    font_col  = ["FFFFFF", "000000", "000000", "FFFFFF", "FFFFFF"]
    for ci, (lbl, bcor, fcor) in enumerate(zip(hdrs, fills_hdr, font_col), start=1):
        _write(ws, row, ci, lbl, bold=True, size=10, fill_hex=bcor, color=fcor)

    statuses = sorted(df["ConfigDispositivo"].unique())
    total_geral = len(df)
    for status in statuses:
        row += 1
        sub = df[df["ConfigDispositivo"] == status]
        dpp  = len(sub[sub["Categoria"] == "Vítima (DPP)"])
        pris = len(sub[sub["Categoria"] == "Sistema Prisional"])
        tot  = len(sub)
        pct  = f"{tot/total_geral*100:.1f}%" if total_geral else "0%"
        fill = AZUL_CLARO if row % 2 == 0 else BRANCO
        _write(ws, row, 1, status, fill_hex=fill, h_align="left", size=10)
        _write(ws, row, 2, dpp,    fill_hex=fill, size=10)
        _write(ws, row, 3, pris,   fill_hex=fill, size=10)
        _write(ws, row, 4, tot,    fill_hex=fill, size=10)
        _write(ws, row, 5, pct,    fill_hex=fill, size=10)

    row += 1
    fill = "D6E4F0"
    dpp_t  = len(df[df["Categoria"] == "Vítima (DPP)"])
    pris_t = len(df[df["Categoria"] == "Sistema Prisional"])
    _write(ws, row, 1, "TOTAL GERAL", bold=True, fill_hex=fill, h_align="left", size=10)
    _write(ws, row, 2, dpp_t,         bold=True, fill_hex=fill, size=10)
    _write(ws, row, 3, pris_t,        bold=True, fill_hex=fill, size=10)
    _write(ws, row, 4, total_geral,   bold=True, fill_hex=fill, size=10)
    _write(ws, row, 5, "100%",        bold=True, fill_hex=fill, size=10)

    _set_col_widths(ws, [28, 16, 18, 14, 10])

    # Gráfico de barras por status
    n_status = len(statuses)
    data_ref_chart = Reference(ws, min_col=4, min_row=5, max_row=5 + n_status)
    cats_ref = Reference(ws, min_col=1, min_row=6, max_row=5 + n_status)
    chart = BarChart()
    chart.type     = "bar"   # horizontal
    chart.grouping = "clustered"
    chart.title    = "Monitorados por Status de Dispositivo"
    chart.style    = 10
    chart.width    = 18
    chart.height   = 12
    chart.add_data(data_ref_chart, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, "G5")

# ── Validação cruzada de totais ───────────────────────────────────────────────
def validar_totais(df_ativos, nome_arquivo):
    total = len(df_ativos)
    dpp   = len(df_ativos[df_ativos["Categoria"] == "Vítima (DPP)"])
    pris  = len(df_ativos[df_ativos["Categoria"] == "Sistema Prisional"])
    fem   = len(df_ativos[df_ativos["SexoNorm"] == "Feminino"])
    masc  = len(df_ativos[df_ativos["SexoNorm"] == "Masculino"])

    ok = True
    if dpp + pris != total:
        print(f"   ⚠ VALIDAÇÃO: DPP({dpp}) + Prisional({pris}) ≠ Total({total})")
        ok = False
    if fem + masc != total:
        print(f"   ⚠ VALIDAÇÃO: Feminino({fem}) + Masculino({masc}) ≠ Total({total})")
        ok = False
    if ok:
        print(f"   ✔ Validação OK — {total} ativos, {dpp} DPP, {pris} Prisional, {fem}F/{masc}M")

# ── Ajustar aba "Resumo" como primeira ───────────────────────────────────────
def mover_resumo_para_frente(wb):
    """Move a aba Resumo para a posição 0."""
    if "Resumo" in wb.sheetnames:
        idx = wb.sheetnames.index("Resumo")
        wb.move_sheet("Resumo", offset=-idx)

# ── Principal ─────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(0)

    arquivo_entrada = sys.argv[1]
    if not os.path.isfile(arquivo_entrada):
        sys.exit(f"❌  Arquivo não encontrado: {arquivo_entrada}")

    agora       = datetime.now()
    data_ref    = agora.strftime("%d/%m/%Y %H:%M")
    nome_base   = os.path.splitext(os.path.basename(arquivo_entrada))[0]
    nome_saida  = sys.argv[2] if len(sys.argv) > 2 else f"Estatisticas_Akiles_{agora.strftime('%Y%m%d_%H%M')}.xlsx"

    print(f"\n📊 Estatísticas Akiles — CME/IAPEN-AP")
    print(f"   Entrada : {arquivo_entrada}")
    print(f"   Saída   : {nome_saida}")
    print()

    print("🔍 Lendo arquivo...")
    df_raw = ler_arquivo(arquivo_entrada)

    print("🔧 Normalizando dados...")
    df = normalizar(df_raw)

    print("🔎 Filtrando ativos/pré-ativos...")
    df_ativos = filtrar_ativos(df)
    print(f"   Total na base: {len(df)} | Ativos: {len(df_ativos)}")

    validar_totais(df_ativos, arquivo_entrada)

    print("\n📑 Gerando abas Excel...")
    wb = openpyxl.Workbook()
    # Remove aba padrão
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    aba_resumo(wb, df, df_ativos, data_ref)
    print("   ✔ Resumo")
    aba_panorama(wb, df, df_ativos, data_ref)
    print("   ✔ Panorama Geral")
    aba_medidas(wb, df_ativos, data_ref)
    print("   ✔ Medidas Protetivas")
    aba_por_tipo(wb, df_ativos, data_ref)
    print("   ✔ Por Tipo")
    aba_por_sexo(wb, df_ativos, data_ref)
    print("   ✔ Por Sexo")
    aba_por_municipio(wb, df_ativos, data_ref)
    print("   ✔ Por Município")
    aba_faixa_etaria(wb, df_ativos, data_ref)
    print("   ✔ Faixa Etária")
    aba_por_grupo(wb, df_ativos, data_ref)
    print("   ✔ Por Grupo")
    aba_por_vara(wb, df_ativos, data_ref)
    print("   ✔ Por Vara/Juízo")
    aba_por_operador(wb, df_ativos, data_ref)
    print("   ✔ Por Operador")
    aba_por_periculosidade(wb, df_ativos, data_ref)
    print("   ✔ Por Periculosidade")
    aba_status_dispositivos(wb, df, data_ref)
    print("   ✔ Status Dispositivos")

    mover_resumo_para_frente(wb)

    wb.save(nome_saida)
    print(f"\n✅  Relatório salvo em: {nome_saida}")
    print(f"    {len(wb.sheetnames)} abas geradas: {', '.join(wb.sheetnames)}\n")

if __name__ == "__main__":
    main()
